from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from collections import defaultdict
# data_only=true returns calculated value if formula is applied on that particular cell


def extract_data(
    file_path,
    worksheet_name,
    rowList,
    colList,
):
    workbook = load_workbook(file_path, data_only=True)
    worksheet = workbook[worksheet_name]
    students = []
    rollNo_col = column_index_from_string("B")
    N_value_col = colList

    for row_index in range(rowList[0], rowList[1]):
        # extracts individual cell values
        rollNo = worksheet.cell(row=row_index, column=rollNo_col).value
        # for loop helps getting all the values mentioned in the list above
        N_value = [
            worksheet.cell(row=row_index, column=col).value for col in N_value_col
        ]

        students.append(N_value)  # to insert data in a 2d list
        # print(students)

    return students


def get_worst(students, mapping):
    
    length = max(mapping)
    count = [0]*length
    for i in range(len(mapping)):
        count[mapping[i]-1] += 1
    performance = []
    for i in range(len(students)):

        performance_temp = [0]*length
        for j in range(len(mapping)):
            performance_temp[mapping[j] - 1] += students[i][j]
        for j in range(length):
            performance_temp[j] = round(performance_temp[j] / count[j] * 10, 2)

        performance.append(performance_temp)
        # performance[i][0] = round(sum(students[i][0:3]) / 30 * 100, 2)
        # performance[i][1] = round(sum(students[i][3:7]) / 40 * 100, 2)
        # performance[i][2] = (students[i][7]) * 10


    return performance

def getMapping(self, examPaper):

        context = "From the given test "+examPaper+" extract the CO number from every question"
        context += "And return only the co numbers without (CO) as a string "
        context += "eg:- 1 1 1 2 2 1 2 3 "

        cos = self.GetTest(context)
        map = cos.strip().split(" ")
        mapping = [int(x) for x in map]
        return mapping

def CO_paper_read(filepath):
    read = PdfReader(filepath)
    pages = read.pages
    res = []
    for page in pages:
        text = page.extract_text()
        y = text.split()
        for i, j in enumerate(y):
            if j == "CO1":
                res.append(1)
            elif j == "CO2":
                res.append(2)
            elif j == "CO3":
                res.append(3)
    print(res)
    return res


def get_CO_Cluster(performance):
    student_wise_CO_cluster = {}
    for i, j in enumerate(performance):
        temp2 = []
        for a, b in enumerate(j):
            if b < 65:
                temp2.append(1)
            else:
                temp2.append(0)
        str1 = int(("".join(str(x) for x in temp2)), 2)

        student_wise_CO_cluster[i] = str1

    cluster = defaultdict(list)

    for i in student_wise_CO_cluster:
        match student_wise_CO_cluster[i]:
            case 0:
                cluster[0].append(i)
            case 1:
                cluster[1].append(i)
            case 2:
                cluster[2].append(i)
            case 3:
                cluster[3].append(i)
            case 4:
                cluster[4].append(i)
            case 5:
                cluster[5].append(i)
            case 6:
                cluster[6].append(i)
            case 7:
                cluster[7].append(i)

    CO_cluster = {k: cluster[k] for k in sorted(cluster)}


# for i in CO_cluster:
#     print(f"{i} = {CO_cluster[i]}")
